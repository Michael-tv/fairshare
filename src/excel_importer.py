"""
Excel importer to load data from the existing spreadsheet.
Handles the "Finances 2024 04.xlsx" format.
"""
from decimal import Decimal
from datetime import datetime, date
from typing import Dict, List, Optional, Tuple, TYPE_CHECKING
import pandas as pd
import openpyxl

from models import (
    Person, Income, Expense, FinancialPeriod,
    IncomeType, ExpenseType, ExpenseCategory
)

if TYPE_CHECKING:
    from transaction_classifier import TransactionClassifier


class ExcelImporter:
    """Imports financial data from Excel spreadsheet."""

    def __init__(self, excel_path: str, classifier: 'TransactionClassifier'):
        """
        Initialize importer.

        Args:
            excel_path: Path to Excel file
            classifier: TransactionClassifier for categorization
        """
        self.excel_path = excel_path
        self.classifier = classifier
        self.wb = openpyxl.load_workbook(excel_path)

    def import_from_expense_balance_sheet(
        self,
        sheet_name: str = "Expense balance sheet"
    ) -> Tuple[FinancialPeriod, Person, Person]:
        """
        Import data from the "Expense balance sheet" tab.

        Returns:
            Tuple of (FinancialPeriod, person1, person2)
        """
        # Read the sheet
        df = pd.read_excel(self.excel_path, sheet_name=sheet_name, header=None)

        # Extract the date (row 0, column 1)
        period_date = df.iloc[0, 1]
        if isinstance(period_date, str):
            period_date = datetime.strptime(period_date, "%Y-%m-%d").date()
        elif isinstance(period_date, datetime):
            period_date = period_date.date()
        else:
            period_date = date.today()

        # Create persons - extract names from the income section
        michael = Person(name="Michael")
        jacqui = Person(name="Jacqui")

        # Create financial period
        period = FinancialPeriod(
            period=period_date,
            people=[michael, jacqui]
        )

        # Parse the sheet row by row
        current_section = None

        for idx, row in df.iterrows():
            # Get values
            col_b = str(row[1]) if pd.notna(row[1]) else ""
            col_c = row[2] if pd.notna(row[2]) else None
            col_d = str(row[3]) if pd.notna(row[3]) else ""

            # Skip empty rows
            if not col_b and not col_c:
                continue

            # Detect sections
            if "Income" in col_b and col_c and isinstance(col_c, (int, float)):
                current_section = "Income"
                continue
            elif "Deductions" in col_b or "TAX" in col_b:
                current_section = "Deductions"
                continue
            elif "Fixed Expenses" in col_b:
                current_section = "Fixed Expenses"
                continue
            elif "Insurance" in col_b and not "Life Insurance" in col_b:
                current_section = "Insurance"
                continue
            elif "Utilities" in col_b or "Variable Costs" in col_b:
                current_section = "Utilities"
                continue
            elif "Misc" in col_b or "Variable" in col_b:
                current_section = "Variable"
                continue

            # Parse income
            if current_section == "Income" and col_b.startswith("Salary"):
                amount = self._to_decimal(col_c)
                if amount and amount > 0:
                    person = michael if "Michael" in col_b else jacqui
                    period.add_income(Income(
                        person=person,
                        amount=amount,
                        income_type=IncomeType.SALARY,
                        description=col_b,
                        period=period_date
                    ))

            # Parse expenses
            elif col_c and isinstance(col_c, (int, float)) and col_c > 0:
                amount = self._to_decimal(col_c)

                # Determine who it belongs to
                person = None
                if "Michael" in col_d or "Michael" in col_b:
                    person = michael
                elif "Jacqui" in col_d or "Jacqui" in col_b:
                    person = jacqui

                # Determine expense type
                if current_section == "Deductions" or "Tax" in col_b or "UIF" in col_b:
                    expense_type = ExpenseType.DEDUCTION
                    category = ExpenseCategory.TAX if "Tax" in col_b else ExpenseCategory.UIF
                elif person:
                    expense_type = ExpenseType.INDIVIDUAL
                    # Use classifier to categorize
                    category_str, _ = self.classifier.classify_transaction(
                        col_b, amount, is_shared_account=False
                    )
                    try:
                        category = ExpenseCategory[category_str]
                    except KeyError:
                        category = ExpenseCategory.OTHER
                else:
                    expense_type = ExpenseType.HOUSEHOLD
                    # Use classifier to categorize
                    category_str, _ = self.classifier.classify_transaction(
                        col_b, amount, is_shared_account=True
                    )
                    try:
                        category = ExpenseCategory[category_str]
                    except KeyError:
                        category = ExpenseCategory.OTHER

                # Add expense
                period.add_expense(Expense(
                    description=col_b,
                    amount=amount,
                    category=category,
                    expense_type=expense_type,
                    belongs_to=person if expense_type in [ExpenseType.INDIVIDUAL, ExpenseType.DEDUCTION] else None,
                    paid_by=person if person else None,
                    period=period_date
                ))

        return period, michael, jacqui

    def import_all_sheets(self) -> Dict[str, FinancialPeriod]:
        """
        Import data from all relevant sheets.

        Returns:
            Dictionary mapping sheet names to FinancialPeriods
        """
        periods = {}

        # Import expense balance sheet
        if "Expense balance sheet" in self.wb.sheetnames:
            period, michael, jacqui = self.import_from_expense_balance_sheet()
            periods["Expense balance sheet"] = period

        return periods

    def _to_decimal(self, value) -> Optional[Decimal]:
        """Convert value to Decimal safely."""
        if value is None or (isinstance(value, str) and not value.strip()):
            return None

        try:
            if isinstance(value, str):
                # Remove currency symbols and spaces
                value = value.replace("R", "").replace(" ", "").replace(",", "")
            return Decimal(str(value))
        except (ValueError, TypeError):
            return None

    def get_summary(self) -> str:
        """Get a summary of what can be imported from the Excel file."""
        summary = []
        summary.append(f"Excel file: {self.excel_path}")
        summary.append(f"Sheets available: {', '.join(self.wb.sheetnames)}")

        for sheet_name in self.wb.sheetnames:
            df = pd.read_excel(self.excel_path, sheet_name=sheet_name, header=None)
            summary.append(f"\n{sheet_name}: {df.shape[0]} rows x {df.shape[1]} columns")

        return "\n".join(summary)


def quick_import(excel_path: str, classifier: Optional['TransactionClassifier'] = None) -> Tuple[FinancialPeriod, Person, Person]:
    """
    Quick import function for the main expense balance sheet.

    Args:
        excel_path: Path to Excel file
        classifier: TransactionClassifier for categorization (creates default if not provided)

    Returns:
        Tuple of (FinancialPeriod, person1, person2)
    """
    # Create classifier if not provided
    if classifier is None:
        from transaction_classifier import TransactionClassifier
        from category_manager import CategoryManager
        cat_mgr = CategoryManager()
        classifier = TransactionClassifier(cat_mgr, use_learned=False)

    importer = ExcelImporter(excel_path, classifier)
    return importer.import_from_expense_balance_sheet()
