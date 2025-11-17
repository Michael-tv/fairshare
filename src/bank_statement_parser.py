"""
Bank Statement Parser - Template-Driven System

Parses bank statements (PDF) using YAML-based templates for multi-bank support.
Templates define bank-specific patterns, enabling parsing of any bank's statements
without code changes.
"""

import re
import decimal
import logging
from decimal import Decimal
from datetime import datetime
from pathlib import Path
from typing import List, Tuple, Optional, Dict, Any
from dataclasses import dataclass
import pdfplumber

from src.bank_template import BankTemplate, TemplateRegistry
from src.parser_diagnostics import get_diagnostics_collector, LineParseAttempt

# Configure logger
logger = logging.getLogger(__name__)


@dataclass
class BankTransaction:
    """Represents a single transaction from a bank statement."""
    date: datetime
    description: str
    amount: Decimal
    account_type: str
    is_credit: bool  # True if payment received (credit), False if expense (debit)
    card_last_digits: Optional[str] = None
    location: Optional[str] = None
    raw_line: str = ""

    @property
    def is_payment(self) -> bool:
        """Alias for is_credit for backward compatibility."""
        return self.is_credit

    def __str__(self):
        direction = "CREDIT" if self.is_credit else "DEBIT "
        return f"{self.date.strftime('%Y-%m-%d')} {direction} R{self.amount:>10,.2f} - {self.description}"


@dataclass
class BankStatementSummary:
    """Summary information from bank statement."""
    statement_date: datetime
    opening_balance: Decimal
    closing_balance: Decimal
    total_expenses: Decimal
    total_payments: Decimal
    interest_fees: Decimal
    statement_number: str
    account_number: str
    credit_count: int  # Number of credit transactions from statement
    debit_count: int   # Number of debit transactions from statement


class BankStatementParser:
    """
    Template-driven bank statement parser.

    Parses PDF bank statements using YAML-based templates that define
    bank-specific patterns and extraction rules.
    """

    def __init__(self, pdf_path: Path, template: BankTemplate):
        """
        Initialize parser with PDF and template.

        Args:
            pdf_path: Path to the PDF bank statement
            template: BankTemplate instance with parsing configuration
        """
        self.pdf_path = pdf_path
        self.template = template
        self.transactions: List[BankTransaction] = []
        self.summary: Optional[BankStatementSummary] = None

    @classmethod
    def create(cls, pdf_path: Path, template_name: Optional[str] = None,
               templates_dir: Path = Path("bank_templates")) -> 'BankStatementParser':
        """
        Factory method to create parser with auto-detection or named template.

        Args:
            pdf_path: Path to PDF bank statement
            template_name: Optional template name (auto-detect if None)
            templates_dir: Directory containing YAML templates

        Returns:
            BankStatementParser instance

        Raises:
            ValueError: If template not found or auto-detection fails
        """
        registry = TemplateRegistry(templates_dir)

        if template_name:
            template = registry.get(template_name)
            if not template:
                raise ValueError(f"Template '{template_name}' not found")

            # Validate against auto-detection
            valid, warning = registry.validate_selection(pdf_path, template_name)
            if warning:
                print(f"⚠️  {warning}")
        else:
            # Auto-detect
            template = registry.auto_detect(pdf_path)
            if not template:
                raise ValueError(
                    "Could not auto-detect bank template. "
                    "Please specify a template with --bank-template"
                )
            print(f"✓ Auto-detected: {template.bank_name} {template.account_type}")

        return cls(pdf_path, template)

    def parse(self) -> Tuple[BankStatementSummary, List[BankTransaction]]:
        """
        Parse the PDF bank statement using template patterns.

        Returns:
            Tuple of (summary, transactions)
        """
        # Start diagnostics session
        collector = get_diagnostics_collector()
        template_name = f"{self.template.bank_name} - {self.template.account_type}"
        collector.start_session(str(self.pdf_path), template_name)

        logger.info(f"Starting parse of {self.pdf_path.name}")
        logger.info(f"Using template: {template_name}")

        try:
            text = self._extract_pdf_text()
            logger.info(f"Extracted text from PDF ({len(text)} characters)")

            self.summary = self._parse_summary(text)
            logger.info(f"Parsed summary - Statement date: {self.summary.statement_date.strftime('%Y-%m-%d')}")

            self.transactions = self._parse_transactions(text)
            logger.info(f"Parsed {len(self.transactions)} transactions")

            # Calculate totals for validation
            total_credits = sum(t.amount for t in self.transactions if t.is_credit)
            total_debits = sum(t.amount for t in self.transactions if not t.is_credit)

            logger.info(f"Total credits: R{total_credits:,.2f}, Total debits: R{total_debits:,.2f}")

            # Validate balance if we have summary info
            if self.summary and (self.summary.opening_balance or self.summary.closing_balance):
                self._validate_balance(total_credits, total_debits)

            return self.summary, self.transactions

        except Exception as e:
            logger.error(f"Error parsing statement: {e}", exc_info=True)
            raise
        finally:
            # End diagnostics session
            collector.end_session()

    def _extract_pdf_text(self) -> str:
        """Extract all text from PDF using pdfplumber."""
        try:
            text = ""
            with pdfplumber.open(self.pdf_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
            return text
        except Exception as e:
            raise RuntimeError(f"Error reading PDF: {e}")

    def _parse_summary(self, text: str) -> BankStatementSummary:
        """Parse summary information using template patterns."""
        summary_config = self.template.config.get('summary', {})

        def extract_field(field_name: str, converter=str, default=None):
            """Extract a field using template pattern."""
            field_cfg = summary_config.get(field_name)
            if not field_cfg:
                return default

            # Handle both single pattern and multiple patterns
            patterns = field_cfg.get('patterns', [])
            if not patterns:
                pattern = field_cfg.get('pattern')
                if pattern:
                    patterns = [pattern]

            for pattern in patterns:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    # Extract value from named group or first group
                    groups = match.groupdict()
                    if 'amount' in groups:
                        value = groups['amount']
                    elif 'date' in groups:
                        value = groups['date']
                    elif 'number' in groups:
                        value = groups['number']
                    elif 'account' in groups:
                        value = groups['account']
                    else:
                        value = match.group(1) if match.lastindex else match.group(0)

                    # Apply converter
                    if converter == Decimal:
                        return self._parse_amount(value)
                    elif converter == datetime:
                        return self._parse_date_string(value, field_cfg)
                    return value.strip() if value else default

            return default

        return BankStatementSummary(
            statement_date=extract_field('statement_date', datetime) or datetime.now(),
            opening_balance=extract_field('opening_balance', Decimal) or Decimal(0),
            closing_balance=extract_field('closing_balance', Decimal) or Decimal(0),
            total_expenses=extract_field('total_expenses', Decimal) or Decimal(0),
            total_payments=extract_field('total_payments', Decimal) or Decimal(0),
            interest_fees=extract_field('interest_fees', Decimal) or Decimal(0),
            statement_number=extract_field('statement_number') or "",
            account_number=extract_field('account_number') or "",
            credit_count=extract_field('credit_count', int) or 0,
            debit_count=extract_field('debit_count', int) or 0
        )

    def _parse_date_string(self, date_str: str, config: Dict[str, Any]) -> datetime:
        """Parse a date string using format from config."""
        if not date_str:
            return datetime.now()

        date_format = config.get('format', '%d %b %Y')

        # Apply month translation if specified
        month_translation = config.get('month_translation', {})
        for afr, eng in month_translation.items():
            date_str = date_str.replace(afr, eng)

        try:
            return datetime.strptime(date_str.strip(), date_format)
        except ValueError as e:
            print(f"⚠️  Could not parse date '{date_str}' with format '{date_format}': {e}")
            return datetime.now()

    def _parse_transactions(self, text: str) -> List[BankTransaction]:
        """Parse transactions using template patterns."""
        collector = get_diagnostics_collector()
        session = collector.current_session

        parsing_cfg = self.template.config.get('parsing', {})
        sections_cfg = self.template.config.get('sections', {})

        # Get transaction pattern
        transaction_pattern = parsing_cfg.get('transaction_pattern')
        if not transaction_pattern:
            logger.error("No transaction_pattern defined in template")
            return []

        # Compile pattern
        try:
            pattern = re.compile(transaction_pattern, re.IGNORECASE)
            logger.debug(f"Compiled transaction pattern: {transaction_pattern[:100]}...")
        except re.error as e:
            logger.error(f"Invalid transaction pattern: {e}")
            return []

        # Section markers
        start_markers = sections_cfg.get('start_markers', [])
        end_markers = sections_cfg.get('end_markers', [])
        skip_lines = sections_cfg.get('skip_lines', [])

        logger.info(f"Section markers - Start: {start_markers}, End: {end_markers}")

        lines = text.split('\n')
        in_section = not start_markers  # If no start markers, parse all lines
        transactions = []
        current_card = None

        # Statistics counters
        line_number = 0
        lines_in_section = 0
        lines_matched = 0
        lines_skipped = 0
        transactions_created = 0
        transactions_failed = 0
        credits_count = 0
        debits_count = 0
        total_credits = Decimal('0')
        total_debits = Decimal('0')
        skip_reasons = {}

        for line in lines:
            line_number += 1
            line_stripped = line.strip()

            # Skip empty lines
            if not line_stripped:
                continue

            # Check for section boundaries
            if start_markers and any(marker in line for marker in start_markers):
                in_section = True
                logger.debug(f"Line {line_number}: Entering transaction section")
                continue

            if end_markers and any(marker in line for marker in end_markers):
                in_section = False
                logger.debug(f"Line {line_number}: Exiting transaction section")
                continue

            # Skip unwanted lines
            skip_reason = None
            for skip_pattern in skip_lines:
                if skip_pattern in line:
                    skip_reason = f"Matches skip pattern: {skip_pattern}"
                    lines_skipped += 1
                    skip_reasons[skip_reason] = skip_reasons.get(skip_reason, 0) + 1
                    break

            if skip_reason:
                if session:
                    attempt = LineParseAttempt(
                        line_number=line_number,
                        line_text=line_stripped[:100],
                        matched=False,
                        transaction_created=False,
                        skip_reason=skip_reason
                    )
                    session.line_attempts.append(attempt)
                continue

            if not in_section:
                continue

            lines_in_section += 1

            # Track card number context
            card_cfg = parsing_cfg.get('card_number', {})
            if card_cfg:
                context_pattern = card_cfg.get('context_pattern')
                if context_pattern:
                    card_match = re.search(context_pattern, line)
                    if card_match:
                        groups = card_match.groupdict()
                        current_card = groups.get('card', card_match.group(1))
                        logger.debug(f"Line {line_number}: Found card context: {current_card}")
                        continue

            # Try to match transaction pattern
            match = pattern.search(line_stripped)

            if match:
                lines_matched += 1

                # Try to parse transaction
                transaction = self._parse_transaction_line(line_stripped, pattern, current_card)

                if transaction:
                    transactions.append(transaction)
                    transactions_created += 1

                    if transaction.is_credit:
                        credits_count += 1
                        total_credits += transaction.amount
                    else:
                        debits_count += 1
                        total_debits += transaction.amount

                    logger.debug(f"Line {line_number}: Parsed transaction - {transaction.description[:30]} R{transaction.amount}")

                    if session:
                        attempt = LineParseAttempt(
                            line_number=line_number,
                            line_text=line_stripped[:100],
                            matched=True,
                            transaction_created=True
                        )
                        session.line_attempts.append(attempt)
                else:
                    transactions_failed += 1
                    error_msg = "Matched pattern but failed validation"
                    logger.warning(f"Line {line_number}: {error_msg} - {line_stripped[:50]}")

                    if session:
                        attempt = LineParseAttempt(
                            line_number=line_number,
                            line_text=line_stripped[:100],
                            matched=True,
                            transaction_created=False,
                            error=error_msg
                        )
                        session.line_attempts.append(attempt)
            else:
                # Line didn't match pattern
                if session:
                    attempt = LineParseAttempt(
                        line_number=line_number,
                        line_text=line_stripped[:100],
                        matched=False,
                        transaction_created=False
                    )
                    session.line_attempts.append(attempt)

        # Update statistics
        if session:
            stats = session.statistics
            stats.total_lines = line_number
            stats.lines_in_section = lines_in_section
            stats.lines_matched_pattern = lines_matched
            stats.lines_skipped = lines_skipped
            stats.transactions_created = transactions_created
            stats.transactions_failed = transactions_failed
            stats.credits_count = credits_count
            stats.debits_count = debits_count
            stats.total_credits = total_credits
            stats.total_debits = total_debits
            stats.skip_reasons = skip_reasons

        logger.info(f"Transaction parsing complete:")
        logger.info(f"  Total lines: {line_number}")
        logger.info(f"  Lines in section: {lines_in_section}")
        logger.info(f"  Lines matched: {lines_matched}")
        logger.info(f"  Transactions created: {transactions_created}")
        logger.info(f"  Transactions failed: {transactions_failed}")

        # Sort by date
        transactions.sort(key=lambda t: t.date)

        return transactions

    def _parse_transaction_line(self, line: str, pattern: re.Pattern,
                                current_card: Optional[str]) -> Optional[BankTransaction]:
        """Parse a single transaction line using template pattern."""
        match = pattern.search(line)
        if not match:
            return None

        groups = match.groupdict()
        parsing_cfg = self.template.config.get('parsing', {})

        try:
            # Parse date
            date = self._parse_transaction_date(groups, parsing_cfg.get('date', {}))

            # Parse amount
            amount = self._parse_transaction_amount(groups, parsing_cfg.get('amount', {}))

            # Determine if credit or debit
            is_credit = self._is_credit_transaction(groups, parsing_cfg.get('amount', {}))

            # Parse description
            description = self._parse_description(groups, parsing_cfg.get('description', {}))

            # Skip if description too short
            min_length = parsing_cfg.get('description', {}).get('min_length', 1)
            if len(description) < min_length:
                return None

            # Extract optional fields
            location = self._extract_location(description, parsing_cfg.get('location', {}))

            # Extract card digits
            card_last_digits = self._extract_card_digits(
                line, current_card, parsing_cfg.get('card_number', {})
            )

            # Get account type from template
            account_type = self.template.config.get('output', {}).get('account_type', 'unknown')

            return BankTransaction(
                date=date,
                description=description,
                amount=amount,
                account_type=account_type,
                is_credit=is_credit,
                card_last_digits=card_last_digits,
                location=location,
                raw_line=line
            )

        except (ValueError, KeyError) as e:
            # Skip lines that don't parse correctly
            return None

    def _parse_transaction_date(self, groups: Dict[str, str],
                               date_cfg: Dict[str, Any]) -> datetime:
        """Parse transaction date from regex groups."""
        day_group = date_cfg.get('day_group', 'day')
        month_group = date_cfg.get('month_group', 'month')
        year_group = date_cfg.get('year_group')

        day_str = groups.get(day_group, '1')
        month_str = groups.get(month_group, 'Jan')

        # Apply month translation (Afrikaans -> English)
        month_translation = date_cfg.get('month_translation', {})
        month_str = month_translation.get(month_str, month_str)

        # Get year
        year_source = date_cfg.get('year_source', 'statement')
        if year_group and groups.get(year_group):
            year = int(groups[year_group])
        elif year_source == 'statement' and self.summary:
            year = self.summary.statement_date.year
        else:
            year = datetime.now().year

        # Parse date
        date_format = date_cfg.get('format', '%d %b %Y')

        # Construct date string
        if year_group:
            date_str = f"{groups.get(year_group)}/{month_str}/{day_str}"
        else:
            date_str = f"{day_str} {month_str} {year}"

        try:
            trans_date = datetime.strptime(date_str, date_format)

            # Adjust year if needed (for year boundary transactions)
            if date_cfg.get('adjust_year') and self.summary:
                if trans_date.month > self.summary.statement_date.month:
                    trans_date = trans_date.replace(year=year - 1)

            return trans_date
        except ValueError:
            # Fallback: try to parse month as number or abbreviation
            try:
                month_num = datetime.strptime(month_str, '%b').month
            except ValueError:
                try:
                    month_num = int(month_str)
                except ValueError:
                    month_num = 1

            day_num = int(day_str) if day_str.isdigit() else 1
            return datetime(year, month_num, day_num)

    def _parse_transaction_amount(self, groups: Dict[str, str],
                                  amount_cfg: Dict[str, Any]) -> Decimal:
        """Parse transaction amount from regex groups."""
        amount_group = amount_cfg.get('group', 'amount')
        amount_str = groups.get(amount_group, '0')

        return self._parse_amount(amount_str, amount_cfg)

    def _parse_amount(self, amount_str: str,
                     config: Optional[Dict[str, Any]] = None) -> Decimal:
        """Parse amount string to Decimal."""
        if not amount_str:
            return Decimal('0')

        # Clean up amount string
        clean = amount_str.strip()

        # Remove currency symbols
        clean = re.sub(r'[R$£€]', '', clean)

        # Handle thousands separators
        if config:
            thousands_seps = config.get('thousands_separators', [])
            if isinstance(thousands_seps, str):
                thousands_seps = [thousands_seps]
            decimal_sep = config.get('decimal_separator', '.')
        else:
            thousands_seps = [',', ' ']
            decimal_sep = '.'

        # Remove thousands separators
        for sep in thousands_seps:
            if sep != decimal_sep:
                clean = clean.replace(sep, '')

        # Remove spaces
        clean = clean.replace(' ', '')

        # Convert decimal separator to period
        if decimal_sep != '.':
            clean = clean.replace(decimal_sep, '.')

        # Remove any remaining non-numeric characters except decimal point and minus
        clean = re.sub(r'[^\d.-]', '', clean)

        try:
            return Decimal(clean)
        except (ValueError, decimal.InvalidOperation):
            return Decimal('0')

    def _is_credit_transaction(self, groups: Dict[str, str],
                               amount_cfg: Dict[str, Any]) -> bool:
        """Determine if transaction is a credit (payment in) or debit (payment out)."""
        credit_cfg = amount_cfg.get('credit_indicator', {})

        if not credit_cfg:
            return False

        indicator_group = credit_cfg.get('group', 'credit')
        credit_value = credit_cfg.get('value', 'Cr')
        debit_value = credit_cfg.get('debit_value', 'Dr')

        indicator = (groups.get(indicator_group) or '').strip()

        # Check if we need to invert logic
        invert = credit_cfg.get('invert', False)

        is_credit = indicator == credit_value

        if invert:
            is_credit = not is_credit

        return is_credit

    def _parse_description(self, groups: Dict[str, str],
                          desc_cfg: Dict[str, Any]) -> str:
        """Parse and clean up description."""
        desc_group = desc_cfg.get('group', 'description')
        description = (groups.get(desc_group) or '').strip()

        # Apply cleanup rules
        cleanup_rules = desc_cfg.get('cleanup', [])
        for rule in cleanup_rules:
            pattern = rule.get('pattern', '')
            replace = rule.get('replace', '')
            description = re.sub(pattern, replace, description)

        return description.strip()

    def _extract_location(self, description: str,
                         loc_cfg: Dict[str, Any]) -> Optional[str]:
        """Extract location code from description."""
        if not loc_cfg:
            return None

        pattern = loc_cfg.get('pattern')
        if not pattern:
            return None

        match = re.search(pattern, description)
        if match:
            loc_group = loc_cfg.get('group', 'country')
            groups = match.groupdict()
            return groups.get(loc_group, match.group(1) if match.lastindex else None)

        return None

    def _extract_card_digits(self, line: str, current_card: Optional[str],
                            card_cfg: Dict[str, Any]) -> Optional[str]:
        """Extract card last digits."""
        if not card_cfg:
            return None

        # Check for inline pattern (within transaction line)
        inline_pattern = card_cfg.get('inline_pattern')
        if inline_pattern:
            match = re.search(inline_pattern, line)
            if match:
                groups = match.groupdict()
                digits = groups.get('digits', match.group(1) if match.lastindex else None)
                if digits:
                    extract_last = card_cfg.get('extract_last_digits', 4)
                    return digits[-extract_last:] if len(digits) >= extract_last else digits

        # Use current card context
        if current_card:
            extract_last = card_cfg.get('extract_last_digits', 4)
            digits = re.sub(r'[^\d]', '', current_card)
            return digits[-extract_last:] if len(digits) >= extract_last else digits

        return None

    def _validate_balance(self, total_credits: Decimal, total_debits: Decimal):
        """Validate parsed transactions against statement balances."""
        collector = get_diagnostics_collector()
        if not collector.current_session or not self.summary:
            return

        opening = self.summary.opening_balance
        closing = self.summary.closing_balance

        # Calculate expected closing balance
        calculated_closing = opening + total_credits - total_debits
        difference = abs(closing - calculated_closing)

        # Store validation results
        validation = {
            'valid': difference < Decimal('0.01'),  # Allow 1 cent rounding
            'opening_balance': opening,
            'closing_balance': closing,
            'calculated_balance': calculated_closing,
            'difference': difference,
            'warnings': []
        }

        if difference >= Decimal('0.01'):
            warning_msg = (
                f"Balance mismatch: Expected R{calculated_closing:,.2f}, "
                f"Statement shows R{closing:,.2f} (Difference: R{difference:,.2f})"
            )
            validation['warnings'].append(warning_msg)
            logger.warning(warning_msg)
        else:
            logger.info(f"Balance validation passed (difference: R{difference:,.2f})")

        # Validate transaction counts if available in statement
        if self.summary.credit_count > 0 or self.summary.debit_count > 0:
            parsed_credits = len([t for t in self.transactions if t.is_credit])
            parsed_debits = len([t for t in self.transactions if not t.is_credit])

            validation['credit_count_expected'] = self.summary.credit_count
            validation['credit_count_parsed'] = parsed_credits
            validation['debit_count_expected'] = self.summary.debit_count
            validation['debit_count_parsed'] = parsed_debits

            if self.summary.credit_count > 0 and parsed_credits != self.summary.credit_count:
                warning_msg = (
                    f"Credit transaction count mismatch: Parsed {parsed_credits}, "
                    f"Statement shows {self.summary.credit_count}"
                )
                validation['warnings'].append(warning_msg)
                logger.warning(warning_msg)

            if self.summary.debit_count > 0 and parsed_debits != self.summary.debit_count:
                warning_msg = (
                    f"Debit transaction count mismatch: Parsed {parsed_debits}, "
                    f"Statement shows {self.summary.debit_count}"
                )
                validation['warnings'].append(warning_msg)
                logger.warning(warning_msg)

            if parsed_credits == self.summary.credit_count and parsed_debits == self.summary.debit_count:
                logger.info(f"Transaction count validation passed: {parsed_credits} credits, {parsed_debits} debits")

        collector.current_session.statistics.balance_validation = validation

    def get_expenses_only(self) -> List[BankTransaction]:
        """Get only expense transactions (exclude payments/credits)."""
        return [t for t in self.transactions if not t.is_credit]

    def get_payments_only(self) -> List[BankTransaction]:
        """Get only payment transactions (credits)."""
        return [t for t in self.transactions if t.is_credit]

    def generate_report(self) -> str:
        """Generate a text report of the statement."""
        lines = []
        lines.append("\n" + "=" * 80)
        lines.append("BANK STATEMENT REPORT")
        lines.append(f"Bank: {self.template.bank_name}")
        lines.append(f"Account Type: {self.template.account_type}")
        lines.append("=" * 80)

        if self.summary:
            lines.append(f"\nStatement Date: {self.summary.statement_date.strftime('%d %b %Y')}")
            lines.append(f"Statement No: {self.summary.statement_number}")
            lines.append(f"Account: {self.summary.account_number}")
            lines.append("")
            lines.append(f"Opening Balance: R{self.summary.opening_balance:>12,.2f}")
            lines.append(f"Total Expenses:  R{self.summary.total_expenses:>12,.2f}")
            lines.append(f"Total Payments:  R{self.summary.total_payments:>12,.2f}")
            if self.summary.interest_fees:
                lines.append(f"Interest/Fees:   R{self.summary.interest_fees:>12,.2f}")
            lines.append(f"Closing Balance: R{self.summary.closing_balance:>12,.2f}")

        lines.append("")
        lines.append("=" * 80)
        lines.append("TRANSACTIONS")
        lines.append("=" * 80)

        expenses = self.get_expenses_only()
        payments = self.get_payments_only()

        if expenses:
            lines.append(f"\nEXPENSES ({len(expenses)} transactions):")
            lines.append("-" * 80)
            lines.append(f"{'Date':<12} {'Description':<50} {'Amount':>12}")
            lines.append("-" * 80)

            for trans in expenses:
                desc = trans.description[:50]
                lines.append(f"{trans.date.strftime('%d %b'):<12} {desc:<50} R{trans.amount:>10,.2f}")

            total = sum(t.amount for t in expenses)
            lines.append("=" * 80)
            lines.append(f"{'TOTAL':<64} R{total:>10,.2f}")
        else:
            lines.append("\nNo expense transactions found.")

        if payments:
            lines.append(f"\nPAYMENTS RECEIVED ({len(payments)} transactions):")
            lines.append("-" * 80)
            lines.append(f"{'Date':<12} {'Description':<50} {'Amount':>12}")
            lines.append("-" * 80)

            for trans in payments:
                desc = trans.description[:50]
                lines.append(f"{trans.date.strftime('%d %b'):<12} {desc:<50} R{trans.amount:>10,.2f}")

            total = sum(t.amount for t in payments)
            lines.append("=" * 80)
            lines.append(f"{'TOTAL':<64} R{total:>10,.2f}")
        else:
            lines.append("\nNo payment transactions found.")

        lines.append("=" * 80)

        return '\n'.join(lines)

    def export_to_excel(self, output_path: Path, person_name: str = "Person"):
        """
        Export transactions to an Excel expense sheet format.

        Args:
            output_path: Path for output Excel file
            person_name: Name to use in sheet (optional)
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("openpyxl required for Excel export. Install with: pip install openpyxl")

        expenses = self.get_expenses_only()
        payments = self.get_payments_only()

        # Create workbook
        wb = Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        # Create Expenses sheet
        ws_expenses = wb.create_sheet('Expenses')

        # Header style
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

        # Write header
        headers = ['Description', 'Amount', 'Category', 'Type']
        for col, header in enumerate(headers, 1):
            cell = ws_expenses.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Write expense transactions (debits)
        for row, trans in enumerate(expenses, 2):
            ws_expenses.cell(row=row, column=1, value=trans.description)
            ws_expenses.cell(row=row, column=2, value=float(trans.amount))
            ws_expenses.cell(row=row, column=3, value='')  # Category - user fills
            ws_expenses.cell(row=row, column=4, value='')  # Type - user fills

        # Write payment transactions (credits) - marked with negative amounts
        # Negative amounts indicate money IN (credits/deposits)
        row_offset = len(expenses) + 2
        for row, trans in enumerate(payments, row_offset):
            ws_expenses.cell(row=row, column=1, value=f"[PAYMENT] {trans.description}")
            ws_expenses.cell(row=row, column=2, value=-float(trans.amount))  # Negative for income
            ws_expenses.cell(row=row, column=3, value='')  # Category - user fills
            ws_expenses.cell(row=row, column=4, value='')  # Type - user fills

        # Adjust column widths
        ws_expenses.column_dimensions['A'].width = 50
        ws_expenses.column_dimensions['B'].width = 15
        ws_expenses.column_dimensions['C'].width = 20
        ws_expenses.column_dimensions['D'].width = 15

        # Create Income sheet (empty template)
        ws_income = wb.create_sheet('Income')

        income_headers = ['Description', 'Amount', 'Type']
        for col, header in enumerate(income_headers, 1):
            cell = ws_income.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        ws_income.column_dimensions['A'].width = 50
        ws_income.column_dimensions['B'].width = 15
        ws_income.column_dimensions['C'].width = 20

        # Save
        wb.save(output_path)

        total_transactions = len(expenses) + len(payments)
        print(f"✓ Exported {len(expenses)} expenses and {len(payments)} payments ({total_transactions} total) to {output_path}")
        return output_path


# Convenience functions for backward compatibility
def parse_bank_statement(pdf_path: str, template_name: Optional[str] = None) -> Tuple[BankStatementSummary, List[BankTransaction]]:
    """
    Convenience function to parse a bank statement.

    Args:
        pdf_path: Path to PDF bank statement
        template_name: Optional template name (auto-detect if None)

    Returns:
        Tuple of (summary, transactions)
    """
    parser = BankStatementParser.create(Path(pdf_path), template_name)
    return parser.parse()


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage: python bank_statement_parser.py <path_to_pdf> [template_name]")
        print("\nExamples:")
        print("  python bank_statement_parser.py statement.pdf")
        print("  python bank_statement_parser.py statement.pdf fnb_credit_card")
        sys.exit(1)

    pdf_path = Path(sys.argv[1])
    template_name = sys.argv[2] if len(sys.argv) > 2 else None

    if not pdf_path.exists():
        print(f"Error: File not found: {pdf_path}")
        sys.exit(1)

    try:
        parser = BankStatementParser.create(pdf_path, template_name)
        summary, transactions = parser.parse()
        print(parser.generate_report())
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)
