"""
Importer for individual person spreadsheets.
Each person maintains their own sheet with their income and expenses.
"""
from decimal import Decimal
from datetime import datetime, date
from typing import Dict, List, Tuple, Optional, TYPE_CHECKING
import pandas as pd
from pathlib import Path

from models import (
    Person, Income, Expense, FinancialPeriod,
    IncomeType, ExpenseType, ExpenseCategory
)

if TYPE_CHECKING:
    from transaction_classifier import TransactionClassifier


class PersonSheetImporter:
    """
    Imports financial data from individual person sheets.

    Expected format:
    - Sheet 1: Income
      Columns: Description, Amount, Type (optional)
      Example:
        Salary, 70000, Salary
        Tax Refund, 5000, Tax Refund
        Rental Income, 8500, Rental

    - Sheet 2: Expenses (Shared Costs Only)
      Columns: Description, Amount, Category (optional)
      Example:
        Groceries at Checkers, 3500, Groceries
        Electricity, 2000, Utilities
        Bond Payment, 14000, Loans
    """

    def __init__(self, classifier: 'TransactionClassifier'):
        """Initialize the importer with a classifier for categorization."""
        self.classifier = classifier
        self.income_keywords = {
            'salary': IncomeType.SALARY,
            'wage': IncomeType.SALARY,
            'rental': IncomeType.RENTAL,
            'rent': IncomeType.RENTAL,
            'business': IncomeType.BUSINESS,
            'investment': IncomeType.INVESTMENT,
            'dividend': IncomeType.INVESTMENT,
            'interest': IncomeType.INVESTMENT,
        }

        self.category_keywords = {
            'groceries': ExpenseCategory.GROCERIES,
            'food': ExpenseCategory.GROCERIES,
            'electricity': ExpenseCategory.UTILITIES,
            'water': ExpenseCategory.UTILITIES,
            'internet': ExpenseCategory.UTILITIES,
            'wifi': ExpenseCategory.UTILITIES,
            'gas': ExpenseCategory.UTILITIES,
            'fuel': ExpenseCategory.FUEL,
            'petrol': ExpenseCategory.FUEL,
            'diesel': ExpenseCategory.FUEL,
            'bond': ExpenseCategory.LOANS,
            'loan': ExpenseCategory.LOANS,
            'mortgage': ExpenseCategory.LOANS,
            'insurance': ExpenseCategory.INSURANCE,
            'medical': ExpenseCategory.MEDICAL_AID,
            'doctor': ExpenseCategory.MEDICAL_AID,
            'pharmacy': ExpenseCategory.MEDICAL_AID,
            'school': ExpenseCategory.SCHOOL_FEES,
            'levies': ExpenseCategory.LEVIES,
            'levy': ExpenseCategory.LEVIES,
            'rates': ExpenseCategory.RATES,
            'cleaning': ExpenseCategory.DOMESTIC_HELP,
            'cleaner': ExpenseCategory.DOMESTIC_HELP,
            'garden': ExpenseCategory.DOMESTIC_HELP,
            'netflix': ExpenseCategory.SUBSCRIPTIONS,
            'dstv': ExpenseCategory.SUBSCRIPTIONS,
            'subscription': ExpenseCategory.SUBSCRIPTIONS,
            'maintenance': ExpenseCategory.MAINTENANCE,
            'repair': ExpenseCategory.MAINTENANCE,
        }

    def import_person_sheet(
        self,
        file_path: str,
        person_name: str,
        period_date: Optional[date] = None,
        income_sheet: str = "Income",
        expenses_sheet: str = "Expenses"
    ) -> Tuple[Person, List[Income], List[Expense]]:
        """
        Import a person's financial sheet.

        Args:
            file_path: Path to Excel file
            person_name: Name of the person
            period_date: Date for this period (defaults to today)
            income_sheet: Name of income sheet
            expenses_sheet: Name of expenses sheet

        Returns:
            Tuple of (Person, list of Incomes, list of Expenses)
        """
        if period_date is None:
            period_date = date.today()

        person = Person(name=person_name)
        incomes = []
        expenses = []

        # Import income
        try:
            income_df = pd.read_excel(file_path, sheet_name=income_sheet)
            incomes = self._parse_income_sheet(income_df, person, period_date)
        except Exception as e:
            print(f"Warning: Could not read income sheet '{income_sheet}': {e}")

        # Import expenses
        try:
            expenses_df = pd.read_excel(file_path, sheet_name=expenses_sheet)
            expenses = self._parse_expenses_sheet(expenses_df, person, period_date)
        except Exception as e:
            print(f"Warning: Could not read expenses sheet '{expenses_sheet}': {e}")

        return person, incomes, expenses

    def _parse_income_sheet(
        self,
        df: pd.DataFrame,
        person: Person,
        period_date: date
    ) -> List[Income]:
        """Parse the income sheet."""
        incomes = []

        # Normalize column names
        df.columns = [str(col).strip().lower() for col in df.columns]

        # Find the relevant columns
        desc_col = self._find_column(df, ['description', 'desc', 'item', 'source'])
        amount_col = self._find_column(df, ['amount', 'value', 'total'])
        type_col = self._find_column(df, ['type', 'category'])

        if desc_col is None or amount_col is None:
            raise ValueError("Could not find Description and Amount columns in income sheet")

        for idx, row in df.iterrows():
            # Skip empty rows
            if pd.isna(row[desc_col]) or pd.isna(row[amount_col]):
                continue

            description = str(row[desc_col]).strip()
            amount = self._parse_amount(row[amount_col])

            if amount <= 0:
                continue

            # Determine income type
            income_type = IncomeType.OTHER
            if type_col and not pd.isna(row[type_col]):
                income_type = self._categorize_income(str(row[type_col]))
            else:
                income_type = self._categorize_income(description)

            incomes.append(Income(
                person=person,
                amount=amount,
                income_type=income_type,
                description=description,
                period=period_date
            ))

        return incomes

    def _parse_expenses_sheet(
        self,
        df: pd.DataFrame,
        person: Person,
        period_date: date
    ) -> List[Expense]:
        """Parse the expenses sheet (shared costs only)."""
        expenses = []

        # Normalize column names
        df.columns = [str(col).strip().lower() for col in df.columns]

        # Find the relevant columns
        desc_col = self._find_column(df, ['description', 'desc', 'item', 'expense'])
        amount_col = self._find_column(df, ['amount', 'value', 'total', 'cost'])
        category_col = self._find_column(df, ['category'])
        type_col = self._find_column(df, ['type', 'expense type', 'split type'])

        if desc_col is None or amount_col is None:
            raise ValueError("Could not find Description and Amount columns in expenses sheet")

        for idx, row in df.iterrows():
            # Skip empty rows
            if pd.isna(row[desc_col]) or pd.isna(row[amount_col]):
                continue

            description = str(row[desc_col]).strip()
            amount = self._parse_amount(row[amount_col])

            if amount <= 0:
                continue

            # Determine expense type (Personal vs Household/Shared)
            expense_type = ExpenseType.HOUSEHOLD  # Default to shared
            if type_col and not pd.isna(row[type_col]):
                type_value = str(row[type_col]).strip().lower()
                # Check if marked as personal
                if type_value in ['personal', 'individual', 'p']:
                    expense_type = ExpenseType.INDIVIDUAL
                # Otherwise treat as shared/household
                elif type_value in ['household', 'shared', 'h', 's']:
                    expense_type = ExpenseType.HOUSEHOLD

            # Determine category using classifier
            if category_col and not pd.isna(row[category_col]):
                # Use provided category if available
                category_str = str(row[category_col]).strip().upper()
                # Try to match to ExpenseCategory
                try:
                    category = ExpenseCategory[category_str]
                except KeyError:
                    # If not a valid enum, use classifier
                    category_str, _ = self.classifier.classify_transaction(
                        description, amount, is_shared_account=False
                    )
                    try:
                        category = ExpenseCategory[category_str]
                    except KeyError:
                        category = ExpenseCategory.OTHER
            else:
                # Use classifier to determine category
                category_str, _ = self.classifier.classify_transaction(
                    description, amount, is_shared_account=False
                )
                try:
                    category = ExpenseCategory[category_str]
                except KeyError:
                    category = ExpenseCategory.OTHER

            # Skip personal expenses from fair share calculation
            # Personal expenses are tracked but don't contribute to the split
            if expense_type == ExpenseType.INDIVIDUAL:
                # Add as individual expense (belongs to this person)
                expenses.append(Expense(
                    description=description,
                    amount=amount,
                    category=category,
                    expense_type=ExpenseType.INDIVIDUAL,
                    belongs_to=person,
                    period=period_date
                ))
            else:
                # Add as shared expense (paid by this person, to be split fairly)
                expenses.append(Expense(
                    description=description,
                    amount=amount,
                    category=category,
                    expense_type=ExpenseType.HOUSEHOLD,
                    paid_by=person,
                    period=period_date
                ))

        return expenses

    def _find_column(self, df: pd.DataFrame, possible_names: List[str]) -> Optional[str]:
        """Find a column by trying multiple possible names."""
        for col in df.columns:
            col_lower = str(col).lower().strip()
            for possible in possible_names:
                if possible.lower() in col_lower:
                    return col
        return None

    def _parse_amount(self, value) -> Decimal:
        """Parse an amount value, handling various formats."""
        if pd.isna(value):
            return Decimal("0")

        # Convert to string and clean
        value_str = str(value).strip()

        # Remove currency symbols and spaces
        value_str = value_str.replace("R", "").replace("$", "").replace(",", "").strip()

        # Handle parentheses as negative
        if value_str.startswith("(") and value_str.endswith(")"):
            value_str = "-" + value_str[1:-1]

        try:
            return Decimal(value_str)
        except:
            return Decimal("0")

    def _categorize_income(self, text: str) -> IncomeType:
        """Categorize income based on description."""
        text_lower = text.lower()

        for keyword, income_type in self.income_keywords.items():
            if keyword in text_lower:
                return income_type

        return IncomeType.OTHER


def import_household_month(
    person1_file: str,
    person1_name: str,
    person2_file: str,
    person2_name: str,
    period_date: Optional[date] = None,
    classifier: Optional['TransactionClassifier'] = None
) -> FinancialPeriod:
    """
    Import a complete household month from two person sheets.

    Args:
        person1_file: Path to person 1's Excel file
        person1_name: Name of person 1
        person2_file: Path to person 2's Excel file
        person2_name: Name of person 2
        period_date: Date for this period (defaults to today)
        classifier: TransactionClassifier for categorization (creates default if not provided)

    Returns:
        Complete FinancialPeriod with all data

    Example:
        >>> period = import_household_month(
        ...     "Michael_April_2024.xlsx", "Michael",
        ...     "Jacqui_April_2024.xlsx", "Jacqui",
        ...     date(2024, 4, 1)
        ... )
    """
    if period_date is None:
        period_date = date.today()

    # Create classifier if not provided
    if classifier is None:
        from transaction_classifier import TransactionClassifier
        from category_manager import CategoryManager
        cat_mgr = CategoryManager()
        classifier = TransactionClassifier(cat_mgr, use_learned=False)

    importer = PersonSheetImporter(classifier)

    # Import person 1
    print(f"Importing {person1_name}'s sheet from {person1_file}...")
    person1, incomes1, expenses1 = importer.import_person_sheet(
        person1_file,
        person1_name,
        period_date
    )

    # Import person 2
    print(f"Importing {person2_name}'s sheet from {person2_file}...")
    person2, incomes2, expenses2 = importer.import_person_sheet(
        person2_file,
        person2_name,
        period_date
    )

    # Create financial period
    period = FinancialPeriod(
        period=period_date,
        people=[person1, person2]
    )

    # Add all income
    for income in incomes1 + incomes2:
        period.add_income(income)

    # Add all expenses
    for expense in expenses1 + expenses2:
        period.add_expense(expense)

    # Summary
    print(f"\nImport Summary:")
    print(f"  Period: {period_date.strftime('%B %Y')}")
    print(f"  {person1_name}: {len(incomes1)} income items, {len(expenses1)} expenses, Total income: R{sum(i.amount for i in incomes1):,.2f}")
    print(f"  {person2_name}: {len(incomes2)} income items, {len(expenses2)} expenses, Total income: R{sum(i.amount for i in incomes2):,.2f}")
    print(f"  Total shared expenses paid by {person1_name}: R{sum(e.amount for e in expenses1):,.2f}")
    print(f"  Total shared expenses paid by {person2_name}: R{sum(e.amount for e in expenses2):,.2f}")

    return period


def create_template_sheets(person_name: str, output_path: str):
    """
    Create template Excel sheets for a person.

    Args:
        person_name: Name of the person
        output_path: Where to save the template

    Example:
        >>> create_template_sheets("Michael", "Michael_Template.xlsx")
    """
    # Create income template
    income_data = {
        'Description': [
            'Monthly Salary',
            'Bonus/Commission',
            'Rental Income',
            'Other Income',
        ],
        'Amount': [
            0.00,
            0.00,
            0.00,
            0.00,
        ],
        'Type': [
            'Salary',
            'Salary',
            'Rental',
            'Other',
        ]
    }

    # Create expenses template
    expenses_data = {
        'Description': [
            'Groceries',
            'Electricity',
            'Water',
            'Internet',
            'Bond/Rent',
            'Insurance',
            'Medical Aid',
            'School Fees',
            'Domestic Help',
            'Fuel',
            'Personal Car Payment',
            'Other',
        ],
        'Amount': [
            0.00,
            0.00,
            0.00,
            0.00,
            0.00,
            0.00,
            0.00,
            0.00,
            0.00,
            0.00,
            0.00,
            0.00,
        ],
        'Category': [
            'Groceries',
            'Utilities',
            'Utilities',
            'Utilities',
            'Loans',
            'Insurance',
            'Medical Aid',
            'School Fees',
            'Domestic Help',
            'Fuel',
            'Loans',
            'Other',
        ],
        'Type': [
            'Household',
            'Household',
            'Household',
            'Household',
            'Household',
            'Household',
            'Household',
            'Household',
            'Household',
            'Household',
            'Personal',
            'Household',
        ]
    }

    income_df = pd.DataFrame(income_data)
    expenses_df = pd.DataFrame(expenses_data)

    # Write to Excel with multiple sheets
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        income_df.to_excel(writer, sheet_name='Income', index=False)
        expenses_df.to_excel(writer, sheet_name='Expenses', index=False)

        # Apply formatting to Expenses sheet
        from openpyxl.styles import PatternFill
        workbook = writer.book
        expenses_sheet = workbook['Expenses']

        # Define fill colors
        household_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
        personal_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")    # Light blue

        # Apply formatting to data rows (skip header row)
        for row_idx, row in enumerate(expenses_df.itertuples(), start=2):  # Start at 2 (row 1 is header)
            expense_type = row.Type

            # Determine which fill to use
            if expense_type and expense_type.lower() in ['household', 'shared', 'h', 's']:
                fill = household_fill
            elif expense_type and expense_type.lower() in ['personal', 'individual', 'p']:
                fill = personal_fill
            else:
                # Default to household color if Type is blank or unrecognized
                fill = household_fill

            # Apply fill to all cells in the row
            for col_idx in range(1, len(expenses_df.columns) + 1):
                cell = expenses_sheet.cell(row=row_idx, column=col_idx)
                cell.fill = fill

        # Add instructions sheet
        instructions = {
            'Instructions': [
                f"Template for {person_name}'s Monthly Finances",
                "",
                "** DEFAULT MODE: NET INCOME (Take-Home Pay) **",
                "",
                "Income Sheet:",
                "- List all income sources for the month",
                "- Use NET amounts (take-home pay from payslip)",
                "- Include salary, bonuses, rental income, tax refunds, etc.",
                "- Tax refunds: Add as 'Other' income type",
                "",
                "NOTE: If you want to use GROSS (before tax) instead,",
                "run with --use-gross flag and the system will calculate tax.",
                "",
                "Expenses Sheet:",
                "- List ALL expenses that YOU paid (both personal and household)",
                "- Use the 'Type' column to mark each expense:",
                "  * 'Household' or 'Shared' = Split fairly with partner",
                "  * 'Personal' or 'Individual' = Your expense only (not split)",
                "",
                "Examples of HOUSEHOLD expenses:",
                "- Groceries, utilities, bond/rent, shared insurance",
                "- Medical aid for family, school fees, domestic help",
                "",
                "Examples of PERSONAL expenses:",
                "- Your personal car payment, gym membership",
                "- Your individual subscriptions, hobbies",
                "",
                "Tips:",
                "- Delete rows you don't need",
                "- Add more rows as needed",
                "- Be as detailed as you like in descriptions",
                "- Categories will be auto-detected if not specified",
                "- Default is 'Household' if Type is blank",
                "",
                "Save as: PersonName_Month_Year.xlsx",
                f"Example: {person_name}_April_2024.xlsx",
            ]
        }

        instructions_df = pd.DataFrame(instructions)
        instructions_df.to_excel(writer, sheet_name='Instructions', index=False, header=False)

    print(f"Template created: {output_path}")
    print(f"Ready for {person_name} to fill in!")
